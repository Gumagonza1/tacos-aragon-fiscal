# -*- coding: utf-8 -*-
"""
excel_export.py - Genera Excel con clasificación correcta de CFDIs.

Soporta DOS modos:
  1. MODO IA (cuando los datos tienen campo 'hoja_ia' de Gemini):
     Hojas: INGRESOS_PLATAFORMAS | GASTOS_DEDUCIBLES | GASTOS_NO_DEDUCIBLES |
            RETENCIONES_APP | NOMINA_EMITIDA
  2. MODO LEGACY (sin IA):
     Hojas: EMITIDOS | RECIBIDOS | RETENCIONES APP

Tasas (Plataformas Tecnológicas, Art. 113-A LISR):
  - IVA emitidos:            16%
  - IVA recibidos:            8%  (zona frontera norte / plataformas)
  - ISR retenido plataformas: 1%  (enajenación de bienes, Art. 113-A frac. III)
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Tasas de impuestos ────────────────────────────────────────────────────────
TASA_IVA_EMITIDO    = 0.16
TASA_IVA_RECIBIDO   = 0.08
TASA_ISR_PLATAFORMA = 0.021  # 2.1% retención plataformas — restaurantes/preparación de alimentos (Art. 113-A frac. II LISR)

# ── Estilos ───────────────────────────────────────────────────────────────────
FONT_HEADER    = Font(name='Arial', bold=True, color='FFFFFF', size=11)
FONT_TITULO    = Font(name='Arial', bold=True, size=14)
FONT_SUBTITULO = Font(name='Arial', bold=True, size=11)
FONT_NORMAL    = Font(name='Arial', size=10)
FONT_TOTAL     = Font(name='Arial', bold=True, size=11)
FONT_RESULTADO = Font(name='Arial', bold=True, size=13)
FONT_META_NOTA = Font(name='Arial', italic=True, size=9, color='666666')

FILL_HEADER    = PatternFill('solid', fgColor='2F5496')
FILL_INGRESO   = PatternFill('solid', fgColor='E2EFDA')
FILL_GASTO     = PatternFill('solid', fgColor='FCE4D6')
FILL_RETENCION = PatternFill('solid', fgColor='D6DCE4')
FILL_NOMINA    = PatternFill('solid', fgColor='EAD1F5')
FILL_RESULTADO = PatternFill('solid', fgColor='FFF2CC')
FILL_PAGAR     = PatternFill('solid', fgColor='F4B084')
FILL_SECCION   = PatternFill('solid', fgColor='D9E2F3')
FILL_ASUNCION  = PatternFill('solid', fgColor='FFFACD')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center')
ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center')

BORDER_THIN       = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'),  bottom=Side(style='thin'))
BORDER_DOUBLE     = Border(top=Side(style='double'), bottom=Side(style='double'))
BORDER_BOTTOM_MED = Border(bottom=Side(style='medium'))

FMT_MONEDA = '$#,##0.00'
FMT_PORCEN = '0.0%'

# ── Columnas de detalle ───────────────────────────────────────────────────────
COLUMNAS = [
    ('fecha',           'Fecha',            14),
    ('tipo_comprobante','Tipo',              6),
    ('emisor_rfc',      'RFC Emisor',        16),
    ('emisor_nombre',   'Emisor',            30),
    ('receptor_rfc',    'RFC Receptor',      16),
    ('receptor_nombre', 'Receptor',          30),
    ('subtotal',        'Subtotal',          15),   # col G
    ('total',           'Total',             15),   # col H
    ('iva_trasladado',  'IVA Trasladado',    15),   # col I
    ('isr_retenido',    'ISR Retenido',      15),   # col J
    ('iva_retenido',    'IVA Retenido',      15),   # col K
    ('uuid',            'UUID',              40),
    ('archivo',         'Archivo',           30),
    ('fuente',          'Fuente',            12),
    ('alerta_ia',       'Alerta IA',         35),
]

COL_IDX = {key: i+1 for i, (key, _, _) in enumerate(COLUMNAS)}

# ── Definición de hojas en MODO IA ────────────────────────────────────────────
# (clave_hoja_ia, titulo_hoja, fill_titulo, tasa_iva, nota_tasa)
_IA_HOJAS = [
    ('INGRESOS_PLATAFORMAS', 'Ingresos de Plataformas (Facturas Emitidas)',    FILL_INGRESO,    TASA_IVA_EMITIDO,  f'IVA {TASA_IVA_EMITIDO:.0%}'),
    ('GASTOS_DEDUCIBLES',    'Gastos Deducibles (Art. 25 y 31 LISR)',          FILL_GASTO,      TASA_IVA_RECIBIDO, f'IVA {TASA_IVA_RECIBIDO:.0%} acreditable'),
    ('GASTOS_NO_DEDUCIBLES', 'Gastos No Deducibles',                           FILL_GASTO,      TASA_IVA_RECIBIDO, f'IVA {TASA_IVA_RECIBIDO:.0%} - NO acreditable'),
    ('RETENCIONES_APP',      'Retenciones de Plataformas (Uber, Didi, Rappi)', FILL_RETENCION,  0,                 'ISR 1% y IVA retenidos por apps'),
    ('NOMINA_EMITIDA',       'Nómina Emitida (Sueldos y Salarios)',             FILL_NOMINA,     0,                 'Deducción de nómina'),
]


# ── Punto de entrada principal ────────────────────────────────────────────────

def generar_excel(datos, rfc_config, nombre=None):
    """
    Genera el Excel fiscal.

    Si los datos tienen campo 'hoja_ia' (clasificación de Gemini) → MODO IA:
      Crea hojas: INGRESOS_PLATAFORMAS, GASTOS_DEDUCIBLES, GASTOS_NO_DEDUCIBLES,
                  RETENCIONES_APP, NOMINA_EMITIDA

    Si no → MODO LEGACY:
      Crea hojas: EMITIDOS, RECIBIDOS, RETENCIONES APP
    """
    if not datos:
        print("  ⚠️  No hay datos para procesar.")
        return

    rfc = rfc_config.strip().upper()

    # Corrección de impuestos (solo IVA trasladado; ISR/IVA retenido se respetan del XML)
    datos = [_asegurar_impuestos(d, rfc) for d in datos]

    n_meta = sum(1 for d in datos if d.get('fuente', '').startswith('metadata'))
    n_xml  = len(datos) - n_meta

    # ¿Tiene clasificación IA?
    tiene_ia = any(d.get('hoja_ia') for d in datos)

    if tiene_ia:
        wb = _construir_workbook_ia(datos, rfc, n_meta, n_xml)
    else:
        wb = _construir_workbook_legacy(datos, rfc, n_meta, n_xml)

    if not nombre:
        nombre = f"Reporte_Fiscal_{rfc}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    os.makedirs("output", exist_ok=True)
    ruta = os.path.join("output", nombre)

    try:
        wb.save(ruta)
        _recalcular(ruta)
        print(f"\n{'='*55}")
        print(f"  ✅ EXCEL GENERADO: {ruta}")
        print(f"  📊 Modo: {'IA (clasificación Gemini)' if tiene_ia else 'Legacy (RFC)'}")
        print(f"  📄 XMLs: {n_xml}  |  Metadata: {n_meta}  |  Total: {len(datos)}")
        print(f"{'='*55}")
        return ruta
    except Exception as e:
        print(f"  ❌ Error al guardar: {e}")
        return None


# ── MODO IA: Construcción del workbook ────────────────────────────────────────

def _construir_workbook_ia(datos, rfc, n_meta, n_xml):
    """Construye el workbook agrupando por hoja_ia de Gemini."""

    # Agrupar datos por hoja_ia; si no tiene, asignar fallback
    grupos = {clave: [] for clave, *_ in _IA_HOJAS}
    ignorados = []

    for d in datos:
        hoja = (d.get('hoja_ia') or '').strip()
        if not hoja:
            hoja = _hoja_fallback(d, rfc)
        if hoja == 'IGNORAR':
            ignorados.append(d)
        elif hoja in grupos:
            grupos[hoja].append(d)
        else:
            # Hoja desconocida → fallback
            grupos[_hoja_fallback(d, rfc)].append(d)

    wb = Workbook()
    primera = True

    # Crear una hoja por cada categoría IA (en orden definido)
    sheet_counts = {}
    for clave, titulo, fill, tasa, nota in _IA_HOJAS:
        registros = grupos.get(clave, [])
        if primera:
            ws = wb.active
            ws.title = clave
            primera = False
        else:
            ws = wb.create_sheet(clave)
        _escribir_detalle(ws, registros, titulo, fill, tasa_iva=tasa, nota_tasa=nota)
        sheet_counts[clave] = len(registros)

    # Hoja de ignorados (si hay)
    if ignorados:
        ws_ign = wb.create_sheet('IGNORADOS')
        _escribir_detalle(ws_ign, ignorados, 'CFDIs Ignorados (sin impacto fiscal)',
                          FILL_ASUNCION, tasa_iva=0, nota_tasa='Sin efecto fiscal')

    # RESUMEN FISCAL
    ws_res = wb.create_sheet('RESUMEN FISCAL')
    _escribir_resumen_ia(ws_res, rfc, sheet_counts, n_meta=n_meta, n_xml=n_xml)

    # Tasas
    ws_cfg = wb.create_sheet('⚙ TASAS')
    _escribir_hoja_tasas(ws_cfg)

    return wb


def _hoja_fallback(d: dict, mi_rfc: str) -> str:
    """Fallback cuando Gemini no clasificó el CFDI."""
    tipo   = d.get('tipo_comprobante', 'I')
    emisor = d.get('emisor_rfc', '').upper()
    if tipo == 'RET':
        return 'RETENCIONES_APP'
    if tipo == 'N' and emisor == mi_rfc:
        return 'NOMINA_EMITIDA'
    if emisor == mi_rfc and tipo in ('I', 'E', 'T'):
        return 'INGRESOS_PLATAFORMAS'
    return 'GASTOS_DEDUCIBLES'


# ── MODO LEGACY: Construcción del workbook ─────────────────────────────────────

def _construir_workbook_legacy(datos, rfc, n_meta, n_xml):
    """Construye el workbook con clasificación clásica (RFC)."""
    emitidos, recibidos, retenciones = [], [], []
    for d in datos:
        tipo   = d.get('tipo_comprobante', '')
        emisor = d.get('emisor_rfc', '').upper()
        if tipo == 'RET':
            retenciones.append(d)
        elif emisor == rfc and tipo in ('I', 'E', 'T', 'N'):
            emitidos.append(d)
        else:
            recibidos.append(d)

    wb = Workbook()
    ws_e = wb.active
    ws_e.title = 'EMITIDOS'
    _escribir_detalle(ws_e, emitidos, 'Facturas Emitidas (Ingresos)', FILL_INGRESO,
                      tasa_iva=TASA_IVA_EMITIDO, nota_tasa='IVA 16% (emitidos)')

    ws_r = wb.create_sheet('RECIBIDOS')
    _escribir_detalle(ws_r, recibidos, 'Facturas Recibidas (Gastos)', FILL_GASTO,
                      tasa_iva=TASA_IVA_RECIBIDO, nota_tasa='IVA 8% (recibidos / plataformas)')

    ws_ret = wb.create_sheet('RETENCIONES APP')
    _escribir_detalle(ws_ret, retenciones, 'Retenciones de Plataformas', FILL_RETENCION,
                      tasa_iva=0, nota_tasa='ISR 1% retenido')

    ws_res = wb.create_sheet('RESUMEN FISCAL')
    _escribir_resumen_legacy(ws_res, rfc, len(emitidos), len(recibidos), len(retenciones),
                             n_meta=n_meta, n_xml=n_xml)

    ws_cfg = wb.create_sheet('⚙ TASAS')
    _escribir_hoja_tasas(ws_cfg)

    return wb


# ── Helpers internos ──────────────────────────────────────────────────────────

def _recalcular(ruta):
    script = "/mnt/skills/public/xlsx/scripts/recalc.py"
    if os.path.exists(script):
        os.system(f"python \"{script}\" \"{ruta}\" 30")


def _asegurar_impuestos(d: dict, mi_rfc: str) -> dict:
    """
    Calcula IVA trasladado para registros de metadata que no lo tienen.

    IMPORTANTE:
      - Los CFDIs tipo RET ya tienen ISR/IVA retenido del XML → no tocar.
      - Las facturas normales (I, E, N) NO tienen ISR/IVA retenido por plataforma;
        eso solo aparece en los CFDI de Retenciones (RET). Evita el doble conteo.
      - Solo completamos 'iva_trasladado' si falta; nunca inventamos retenciones.
    """
    d = dict(d)  # copia para no mutar el original
    tipo   = d.get('tipo_comprobante', 'I')
    fuente = d.get('fuente', 'xml')

    # CFDIs de retención: valores vienen del XML, no tocar
    if tipo == 'RET':
        return d

    subtotal = float(d.get('subtotal') or 0)
    iva_tras = float(d.get('iva_trasladado') or 0)

    # XML con IVA explícito: confiable, no recalcular
    if fuente == 'xml' and iva_tras > 0:
        return d

    # Para registros de metadata: limpiar retenciones calculadas incorrectamente.
    # ISR/IVA retenido son EXCLUSIVOS de los CFDI RET (ya procesados arriba).
    if 'metadata' in fuente:
        d['isr_retenido'] = 0
        d['iva_retenido']  = 0

    # Derivar subtotal del total si falta
    if subtotal == 0 and float(d.get('total') or 0) > 0:
        total      = float(d['total'])
        es_emitido = d.get('emisor_rfc', '').upper() == mi_rfc.upper()
        tasa       = TASA_IVA_EMITIDO if es_emitido else TASA_IVA_RECIBIDO
        subtotal   = total / (1 + tasa)
        d['subtotal'] = round(subtotal, 2)

    # Calcular IVA trasladado si falta
    if subtotal > 0 and iva_tras == 0:
        es_emitido    = d.get('emisor_rfc', '').upper() == mi_rfc.upper()
        tasa_iva      = TASA_IVA_EMITIDO if es_emitido else TASA_IVA_RECIBIDO
        d['iva_trasladado'] = round(subtotal * tasa_iva, 2)
        d.setdefault('fuente', 'metadata_calc')

    return d


# ── Hoja de detalle ───────────────────────────────────────────────────────────

def _escribir_detalle(ws, registros, titulo, fill_titulo, tasa_iva=0.16, nota_tasa=''):
    n = len(registros)

    ws.merge_cells('A1:O1')
    c = ws['A1']
    c.value = titulo
    c.font  = FONT_TITULO
    c.fill  = fill_titulo
    c.alignment = ALIGN_CENTER

    ws.merge_cells('A2:O2')
    texto_sub = f"Total: {n} documentos"
    if nota_tasa:
        texto_sub += f"   |   {nota_tasa}"
    ws['A2'].value     = texto_sub
    ws['A2'].font      = FONT_SUBTITULO
    ws['A2'].alignment = ALIGN_CENTER
    ws['A2'].fill      = FILL_ASUNCION

    # Headers (fila 4)
    for ci, (_, label, width) in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=4, column=ci, value=label)
        cell.font      = FONT_HEADER
        cell.fill      = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER_THIN
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Datos (fila 5+)
    CAMPOS_MONEDA = {'subtotal', 'total', 'iva_trasladado', 'isr_retenido', 'iva_retenido'}
    for ri, reg in enumerate(registros, 5):
        es_meta = reg.get('fuente', 'xml') not in ('xml', 'xml_ret_v1', 'xml_ret_v2')
        for ci, (key, _, _) in enumerate(COLUMNAS, 1):
            val  = reg.get(key, '')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font   = FONT_NORMAL
            cell.border = BORDER_THIN
            if key in CAMPOS_MONEDA:
                cell.number_format = FMT_MONEDA
                cell.alignment     = ALIGN_RIGHT
                if es_meta:
                    cell.fill = FILL_ASUNCION
            if key == 'fuente' and es_meta:
                cell.font = FONT_META_NOTA
            if key == 'alerta_ia' and val:
                cell.font = Font(name='Arial', size=9, color='C00000')

    # Fila de totales con SUM
    if registros:
        tr = 5 + n
        ws.cell(row=tr, column=1, value='TOTALES').font = FONT_TOTAL
        for ci, (key, _, _) in enumerate(COLUMNAS, 1):
            if key in CAMPOS_MONEDA:
                cl   = get_column_letter(ci)
                cell = ws.cell(row=tr, column=ci)
                cell.value         = f'=SUM({cl}5:{cl}{tr - 1})'
                cell.font          = FONT_TOTAL
                cell.number_format = FMT_MONEDA
                cell.border        = BORDER_DOUBLE

    ws.auto_filter.ref = f"A4:O{4 + n}"
    ws.freeze_panes    = 'A5'


# ── Hoja de tasas ─────────────────────────────────────────────────────────────

def _escribir_hoja_tasas(ws):
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15

    ws['A1'] = '⚙️ TASAS DE IMPUESTOS APLICADAS'
    ws['A1'].font = Font(name='Arial', bold=True, size=13, color='2F5496')

    ws['A2'] = 'Régimen: Plataformas Tecnológicas (Art. 113-A LISR) | RMF 2026'
    ws['A2'].font = FONT_SUBTITULO

    filas = [
        ('Tasa IVA — Facturas EMITIDAS (ingresos propios)',          TASA_IVA_EMITIDO),
        ('Tasa IVA — Facturas RECIBIDAS (gastos, zona frontera)',     TASA_IVA_RECIBIDO),
        ('Tasa ISR — Retención por plataformas (Art. 113-A frac III, enajenación de bienes)', TASA_ISR_PLATAFORMA),
    ]
    for i, (label, tasa) in enumerate(filas, 4):
        ws.cell(row=i, column=1, value=label).font = FONT_NORMAL
        c = ws.cell(row=i, column=2, value=tasa)
        c.number_format = FMT_PORCEN
        c.font          = Font(name='Arial', bold=True, size=11, color='0070C0')
        c.alignment     = ALIGN_CENTER

    ws['A8']  = '⚠️ Los valores con fondo amarillo son ESTIMADOS (calculados desde Metadata del SAT).'
    ws['A8'].font  = Font(name='Arial', italic=True, size=10, color='C00000')
    ws['A9']  = '   Los XMLs parseados tienen valores exactos del comprobante fiscal.'
    ws['A9'].font  = FONT_META_NOTA
    ws['A10'] = '   ISR retenido = 1% (enajenación bienes plataformas). NO aplica a facturas normales.'
    ws['A10'].font = FONT_META_NOTA


# ── RESUMEN FISCAL — MODO IA ──────────────────────────────────────────────────

def _escribir_resumen_ia(ws, rfc, sheet_counts: dict, n_meta=0, n_xml=0):
    """
    Genera el RESUMEN FISCAL referenciando las hojas clasificadas por Gemini.

    sheet_counts: dict con {nombre_hoja: n_registros}
    """
    ws.sheet_properties.tabColor = 'FF6600'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 58
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 4

    # Número de la fila de totales en cada hoja
    def tr(hoja):
        n = sheet_counts.get(hoja, 0)
        return 5 + n  # fila TOTALES

    # Referencia segura a totales de una hoja
    def ref(hoja, col):
        n = sheet_counts.get(hoja, 0)
        if n == 0:
            return 0
        return f"='{hoja}'!{col}{tr(hoja)}"

    n_inp  = sheet_counts.get('INGRESOS_PLATAFORMAS', 0)
    n_gd   = sheet_counts.get('GASTOS_DEDUCIBLES', 0)
    n_gnd  = sheet_counts.get('GASTOS_NO_DEDUCIBLES', 0)
    n_ret  = sheet_counts.get('RETENCIONES_APP', 0)
    n_nom  = sheet_counts.get('NOMINA_EMITIDA', 0)

    # ── Encabezado ──────────────────────────────────────────────────────────
    ws.merge_cells('B1:C1')
    _cell(ws, 'B1', '🌮 RESUMEN FISCAL — TACOS ARAGÓN (Clasificación IA)',
          Font(name='Arial', bold=True, size=15, color='2F5496'), alignment=ALIGN_CENTER)
    ws.merge_cells('B2:C2')
    _cell(ws, 'B2', f'RFC: {rfc}',
          Font(name='Arial', size=11, color='666666'), alignment=ALIGN_CENTER)
    ws.merge_cells('B3:C3')
    _cell(ws, 'B3',
          f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  XMLs: {n_xml}  |  Metadata: {n_meta}',
          Font(name='Arial', size=9, color='999999'), alignment=ALIGN_CENTER)

    # ── Sección 1: Ingresos de plataformas ──────────────────────────────────
    _seccion(ws, 5, f'1. INGRESOS DE PLATAFORMAS — IVA {TASA_IVA_EMITIDO:.0%}')
    _fila(ws, 6,  'Subtotal ingresos (sin IVA)',               ref('INGRESOS_PLATAFORMAS', 'G'))
    _fila(ws, 7,  f'IVA cobrado ({TASA_IVA_EMITIDO:.0%})',     ref('INGRESOS_PLATAFORMAS', 'I'))
    _fila(ws, 8,  'Total facturado (con IVA)',                  ref('INGRESOS_PLATAFORMAS', 'H'))
    _fila(ws, 9,  'Número de facturas emitidas',                n_inp, es_entero=True)

    # ── Sección 2: Gastos deducibles ─────────────────────────────────────────
    _seccion(ws, 11, f'2. GASTOS DEDUCIBLES — IVA {TASA_IVA_RECIBIDO:.0%} acreditable')
    _fila(ws, 12, 'Subtotal gastos deducibles',                 ref('GASTOS_DEDUCIBLES', 'G'))
    _fila(ws, 13, f'IVA acreditable ({TASA_IVA_RECIBIDO:.0%})',ref('GASTOS_DEDUCIBLES', 'I'))
    _fila(ws, 14, 'Total pagado con IVA',                       ref('GASTOS_DEDUCIBLES', 'H'))
    _fila(ws, 15, 'Número de facturas deducibles',              n_gd, es_entero=True)

    # ── Sección 3: Gastos no deducibles ─────────────────────────────────────
    _seccion(ws, 17, '3. GASTOS NO DEDUCIBLES (referencia, no acreditan IVA)')
    _fila(ws, 18, 'Total gastos no deducibles',                 ref('GASTOS_NO_DEDUCIBLES', 'H'))
    _fila(ws, 19, 'IVA pagado (NO acreditable)',                ref('GASTOS_NO_DEDUCIBLES', 'I'))
    _fila(ws, 20, 'Número de facturas',                         n_gnd, es_entero=True)

    # ── Sección 4: Retenciones de plataformas ────────────────────────────────
    _seccion(ws, 22, f'4. RETENCIONES DE PLATAFORMAS — ISR {TASA_ISR_PLATAFORMA:.0%} (Art. 113-A LISR)')
    _fila(ws, 23, f'ISR retenido por apps ({TASA_ISR_PLATAFORMA:.0%})',  ref('RETENCIONES_APP', 'J'))
    _fila(ws, 24, 'IVA retenido por apps (50% IVA)',            ref('RETENCIONES_APP', 'K'))
    _fila(ws, 25, 'Monto total base gravable (certificados)',   ref('RETENCIONES_APP', 'H'))
    _fila(ws, 26, 'Número de certificados de retención',        n_ret, es_entero=True)

    # ── Sección 5: Nómina ────────────────────────────────────────────────────
    _seccion(ws, 28, '5. NÓMINA EMITIDA (Sueldos y Salarios — deducible)')
    _fila(ws, 29, 'Total nómina emitida',                       ref('NOMINA_EMITIDA', 'H'))
    _fila(ws, 30, 'Número de recibos de nómina',               n_nom, es_entero=True)

    # ── Sección 6: Cálculo IVA ───────────────────────────────────────────────
    _seccion(ws, 32, f'📊 CÁLCULO DE IVA DEL PERIODO', fill=FILL_RESULTADO)
    _fila(ws, 33, f'(+) IVA cobrado {TASA_IVA_EMITIDO:.0%} (ingresos plataformas)',    '=C7')
    _fila(ws, 34, f'(-) IVA acreditable {TASA_IVA_RECIBIDO:.0%} (gastos deducibles)',  '=C13')
    _fila(ws, 35, '(-) IVA retenido por plataformas (apps)',                             '=C24')
    _fila_resultado(ws, 36, 'IVA A PAGAR  (negativo = saldo a favor)', '=C33-C34-C35')

    # ── Sección 7: Cálculo ISR ───────────────────────────────────────────────
    _seccion(ws, 38, f'📊 CÁLCULO DE ISR PROVISIONAL — Pagos Provisionales', fill=FILL_RESULTADO)
    _fila(ws, 39, '(+) Ingresos brutos plataformas (base gravable certificados)', '=C25')
    _fila(ws, 40, '(+) Otros ingresos (facturas emitidas)',                        '=C6')
    _fila(ws, 41, '(-) Gastos deducibles (subtotal)',                              '=C12')
    _fila(ws, 42, '(-) Nómina emitida deducible',                                  '=C29')
    _fila(ws, 43, '(=) Base gravable ISR',                                         '=C39+C40-C41-C42', bold=True)
    _fila(ws, 44, f'(-) ISR retenido por plataformas ({TASA_ISR_PLATAFORMA:.0%})', '=C23')
    _fila_resultado(ws, 45, 'ISR PROVISIONAL A PAGAR  (negativo = saldo a favor)', '=C43-C44')

    _seccion(ws, 47, '⚠️  NOTA: Aplicar tarifa mensual Art. 96 LISR o tabla RESICO sobre la base (C43)',
             fill=PatternFill('solid', fgColor='FFF2CC'))
    ws[f'B47'].font = Font(name='Arial', italic=True, size=9, color='C00000')

    # ── Sección 8: Total estimado a pagar ───────────────────────────────────
    _seccion(ws, 49, '💰  TOTAL ESTIMADO A PAGAR AL SAT', fill=FILL_PAGAR)
    _fila(ws, 50, 'IVA a pagar (si C36 > 0)',              '=C36')
    _fila(ws, 51, 'ISR provisional a pagar (si C45 > 0)',  '=C45')

    ws['B52'].value = '═══  TOTAL ESTIMADO A PAGAR'
    ws['B52'].font  = Font(name='Arial', bold=True, size=14, color='C00000')
    ws['B52'].border = BORDER_THIN
    ws['C52'].value  = '=MAX(0,C50)+MAX(0,C51)'
    ws['C52'].font   = Font(name='Arial', bold=True, size=14, color='C00000')
    ws['C52'].number_format = FMT_MONEDA
    ws['C52'].fill   = FILL_PAGAR
    ws['C52'].border = BORDER_DOUBLE
    ws['C52'].alignment = ALIGN_RIGHT

    _fila(ws, 54, 'IVA a favor (si C36 < 0)',  '=IF(C36<0,ABS(C36),0)')
    _fila(ws, 55, 'ISR a favor (si C45 < 0)',  '=IF(C45<0,ABS(C45),0)')

    # ── Notas ────────────────────────────────────────────────────────────────
    ws['B57'].value = '⚠️ NOTAS IMPORTANTES:'
    ws['B57'].font  = Font(name='Arial', bold=True, size=10, color='C00000')
    notas = [
        f'• Clasificación por Gemini IA según RMF 2026, Art. 113-A y 25 LISR.',
        f'• IVA emitidos {TASA_IVA_EMITIDO:.0%} | IVA recibidos {TASA_IVA_RECIBIDO:.0%} | ISR retenido plataformas {TASA_ISR_PLATAFORMA:.0%}.',
        '• ISR retenido es el anticipo que ya pagaron las apps (acreditable 100%).',
        '• IVA retenido por apps = 50% del IVA cobrado (acreditable).',
        '• Solo GASTOS_DEDUCIBLES acreditan IVA; GASTOS_NO_DEDUCIBLES no.',
        '• Celdas en amarillo = impuestos estimados (Metadata, no XML).',
        '• La base ISR real requiere tarifa mensual. Consulta a tu contador.',
        '• Este cálculo es estimado. Confirma con contador para declaración definitiva.',
    ]
    for i, nota in enumerate(notas, 58):
        ws[f'B{i}'].value = nota
        ws[f'B{i}'].font  = Font(name='Arial', size=9, color='666666')

    ws.sheet_view.showGridLines = False


# ── RESUMEN FISCAL — MODO LEGACY ──────────────────────────────────────────────

def _escribir_resumen_legacy(ws, rfc, n_e, n_r, n_t, n_meta=0, n_xml=0):
    ws.sheet_properties.tabColor = 'FF6600'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 4

    te = 5 + n_e
    tr = 5 + n_r
    tt = 5 + n_t

    ws.merge_cells('B1:C1')
    _cell(ws, 'B1', '🌮 RESUMEN FISCAL - TACOS ARAGÓN',
          Font(name='Arial', bold=True, size=16, color='2F5496'), alignment=ALIGN_CENTER)
    ws.merge_cells('B2:C2')
    _cell(ws, 'B2', f'RFC: {rfc}',
          Font(name='Arial', size=11, color='666666'), alignment=ALIGN_CENTER)
    ws.merge_cells('B3:C3')
    _cell(ws, 'B3', f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  XMLs: {n_xml}  |  Metadata: {n_meta}',
          Font(name='Arial', size=9, color='999999'), alignment=ALIGN_CENTER)

    _seccion(ws, 5, f'1. INGRESOS — IVA {TASA_IVA_EMITIDO:.0%} (FACTURAS EMITIDAS)')
    _fila(ws, 6,  'Total Facturado (subtotal)',         f"='EMITIDOS'!G{te}" if n_e else 0)
    _fila(ws, 7,  f'IVA Cobrado ({TASA_IVA_EMITIDO:.0%})', f"='EMITIDOS'!I{te}" if n_e else 0)
    _fila(ws, 8,  'Total con IVA',                      f"='EMITIDOS'!H{te}" if n_e else 0)
    _fila(ws, 9,  'Número de facturas emitidas',         n_e, es_entero=True)

    _seccion(ws, 11, f'2. GASTOS — IVA {TASA_IVA_RECIBIDO:.0%} (FACTURAS RECIBIDAS)')
    _fila(ws, 12, 'Total Gastos (subtotal)',             f"='RECIBIDOS'!G{tr}" if n_r else 0)
    _fila(ws, 13, f'IVA Pagado / Acreditable ({TASA_IVA_RECIBIDO:.0%})', f"='RECIBIDOS'!I{tr}" if n_r else 0)
    _fila(ws, 14, 'Total pagado con IVA',                f"='RECIBIDOS'!H{tr}" if n_r else 0)
    _fila(ws, 15, 'Número de facturas recibidas',        n_r, es_entero=True)

    _seccion(ws, 17, f'3. RETENCIONES — ISR {TASA_ISR_PLATAFORMA:.0%} (CERTIFICADOS PLATAFORMAS)')
    _fila(ws, 18, f'ISR Retenido por plataformas ({TASA_ISR_PLATAFORMA:.0%})', f"='RETENCIONES APP'!J{tt}" if n_t else 0)
    _fila(ws, 19, 'IVA Retenido por plataformas',       f"='RETENCIONES APP'!K{tt}" if n_t else 0)
    _fila(ws, 20, 'Total retenido',                     '=C18+C19', bold=True)
    _fila(ws, 21, 'Número de certificados',              n_t, es_entero=True)

    _seccion(ws, 23, f'📊 CÁLCULO DE IVA', fill=FILL_RESULTADO)
    _fila(ws, 24, f'(+) IVA Cobrado {TASA_IVA_EMITIDO:.0%}',         '=C7')
    _fila(ws, 25, f'(-) IVA Acreditable {TASA_IVA_RECIBIDO:.0%}',    '=C13')
    _fila(ws, 26, '(-) IVA Retenido por plataformas',                  '=C19')
    _fila_resultado(ws, 27, 'IVA A PAGAR  (negativo = a favor)', '=C24-C25-C26')

    _seccion(ws, 29, f'📊 CÁLCULO ISR PROVISIONAL', fill=FILL_RESULTADO)
    _fila(ws, 30, '(+) Ingresos del periodo',           '=C6')
    _fila(ws, 31, '(-) Gastos deducibles',              '=C12')
    _fila(ws, 32, '(=) Utilidad estimada',              '=C30-C31', bold=True)
    _fila(ws, 33, f'(-) ISR retenido plataformas ({TASA_ISR_PLATAFORMA:.0%})', '=C18')
    _fila_resultado(ws, 34, 'ISR PROVISIONAL A PAGAR  (negativo = a favor)', '=C32-C33')

    _seccion(ws, 36, '💰  TOTAL ESTIMADO A PAGAR AL SAT', fill=FILL_PAGAR)
    _fila(ws, 37, 'IVA a pagar',  '=C27')
    _fila(ws, 38, 'ISR a pagar',  '=C34')
    ws['B39'].value = '═══  TOTAL ESTIMADO'
    ws['B39'].font  = Font(name='Arial', bold=True, size=14, color='C00000')
    ws['C39'].value = '=MAX(0,C37)+MAX(0,C38)'
    ws['C39'].font  = Font(name='Arial', bold=True, size=14, color='C00000')
    ws['C39'].number_format = FMT_MONEDA
    ws['C39'].fill   = FILL_PAGAR
    ws['C39'].border = BORDER_DOUBLE
    ws['C39'].alignment = ALIGN_RIGHT

    ws.sheet_view.showGridLines = False


# ── Helpers de celda ─────────────────────────────────────────────────────────

def _cell(ws, ref, value, font=None, fill=None, alignment=None, fmt=None):
    c = ws[ref]
    c.value = value
    if font:      c.font = font
    if fill:      c.fill = fill
    if alignment: c.alignment = alignment
    if fmt:       c.number_format = fmt
    return c


def _seccion(ws, row, texto, fill=None):
    ws.merge_cells(f'B{row}:C{row}')
    c = ws[f'B{row}']
    c.value     = texto
    c.font      = Font(name='Arial', bold=True, size=12, color='2F5496')
    c.fill      = fill or FILL_SECCION
    c.alignment = ALIGN_LEFT
    c.border    = BORDER_THIN
    ws[f'C{row}'].border = BORDER_THIN


def _fila(ws, row, label, valor, bold=False, es_entero=False):
    b, c = ws[f'B{row}'], ws[f'C{row}']
    b.value     = label
    b.font      = FONT_TOTAL if bold else FONT_NORMAL
    b.alignment = ALIGN_LEFT
    b.border    = BORDER_THIN
    c.value     = valor
    c.font      = FONT_TOTAL if bold else FONT_NORMAL
    c.number_format = '#,##0' if es_entero else FMT_MONEDA
    c.alignment = ALIGN_RIGHT
    c.border    = BORDER_THIN if not bold else BORDER_BOTTOM_MED


def _fila_resultado(ws, row, label, formula):
    b, c = ws[f'B{row}'], ws[f'C{row}']
    b.value  = f'═══  {label}'
    b.font   = FONT_RESULTADO
    b.border = BORDER_THIN
    c.value  = formula
    c.font   = FONT_RESULTADO
    c.number_format = FMT_MONEDA
    c.fill   = FILL_RESULTADO
    c.alignment = ALIGN_RIGHT
    c.border = BORDER_DOUBLE
